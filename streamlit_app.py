import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime, re, io

st.set_page_config(page_title='Route Processor', page_icon='🚛', layout='centered')
st.markdown("""
<style>
.stButton>button { background-color: #1D3557; color: white; width: 100%; padding: 0.6rem; font-size: 1.1rem; border-radius: 8px; border: none; }
.stButton>button:hover { background-color: #457B9D; }
</style>""", unsafe_allow_html=True)

st.markdown("<h1 style='color:#1D3557'>🚛 Route Processor</h1>", unsafe_allow_html=True)
st.markdown("<p style='color:#666'>Upload your route CSV — Randwick, Burwood, Woollahra and Bayside supported.</p>", unsafe_allow_html=True)
st.divider()

NAVY='1D3557'; MID_BLUE='457B9D'; LIGHT_BLUE='A8D8EA'; ALT_ROW='EBF5FB'
WHITE='FFFFFF'; DARK_TEXT='1A1A2E'; TOTAL_BG='D4E6F1'
ORANGE='C75B00'; LIGHT_ORANGE='FEF0E7'; ORANGE_MID='E8874A'; ORANGE_HDR='FDDCCA'
GRN_DARK='1B4332'; GRN_MID='2D6A4F'; GRN_LIGHT='D8F3DC'; GRN_ALT='F0FBF3'

def fill(hex): return PatternFill('solid', fgColor=hex)
def font(bold=False, sz=10, color=DARK_TEXT, italic=False):
    return Font(name='Calibri', bold=bold, size=sz, color=color, italic=italic)
def thin_border(color='D5E8F5'): return Border(bottom=Side(style='thin', color=color))

COL_CAPS = {
    'Ref ID':10,'Date':12,'Date Booked':12,'Date Collected':14,
    'Suburb':14,'Qty Booked':9,'Qty Collected':11,'Collected':11,
    'Collection Notes':45,'Notes':45,'Tracking Link':40,
}
def col_width(series, header, cap=None):
    longest = max(series.astype(str).str.len().max(), len(header)) + 2
    c = cap or COL_CAPS.get(header, None)
    return int(min(longest, c) if c else longest)

def clean_date(val):
    val = str(val).strip().split(' ')[0]
    try: return pd.to_datetime(val, dayfirst=True)
    except: return pd.NaT

def extract_suburb_from_address(address):
    address = str(address)
    m = re.search(r',\s*([A-Za-z\s]+),\s*(?:Randwick|Burwood|Woollahra Municipal|Bayside)\s*Council', address, re.IGNORECASE)
    if m: return m.group(1).strip().title()
    m = re.search(r'\b([A-Z]+(?:\s+[A-Z]+)?)\s*,\s*NSW', address)
    if m: return m.group(1).strip().title()
    m = re.search(r',\s*([A-Z][A-Z\s]+)\s*$', address, re.IGNORECASE)
    if m: return m.group(1).strip().title()
    return ''

def extract_suburb(row):
    if pd.isna(row['suburb']) or str(row['suburb']).strip() == '':
        return extract_suburb_from_address(row['address'])
    return str(row['suburb']).strip().title()

def detect_council(df):
    route = str(df['route'].iloc[0]).lower()
    if 'burwood' in route: return 'burwood'
    if 'woollahra' in route: return 'woollahra'
    if 'bayside' in route: return 'bayside'
    if 'innerwest' in route: return 'innerwest'
    if 'penrith' in route: return 'penrith'
    if 'randwick' in route: return 'randwick'
    return 'unknown'

col_map_oncall = {
    'seller_order_id':'Ref ID','date_booked':'Date','suburb':'Suburb',
    'address':'Address','notes':'Collection Notes','qty_booked':'Qty Booked',
    'driver_provided_recipient_notes':'Qty Collected',
    'photo_url':'Photo URL','tracking_url':'Tracking Link',
}
col_map_booked = {
    'seller_order_id':'Ref ID','date_booked':'Date Booked',
    'address':'Address','notes':'Notes',
    'driver_provided_recipient_notes':'Qty Collected',
    'photo_url':'Photo URL','tracking_url':'Tracking Link',
}
col_map_burwood = {
    'address':'Address',
    'driver_provided_internal_notes':'Collected',
    'photo_url':'Photo URL',
}

def build_oncall(ws, df_data, route_name):
    df_out = df_data[list(col_map_oncall.keys())].rename(columns=col_map_oncall)
    headers = list(col_map_oncall.values()); num_cols = len(headers)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    c = ws['A1']; c.value = f'  {route_name}  —  On-Call Mattress Collection'
    c.fill = fill(NAVY); c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.date.today().strftime("%d/%m/%Y")}   |   '
               f'Total stops: {len(df_out)}   |   Qty collected: {int(df_out["Qty Collected"].sum())}')
    c.fill = fill(MID_BLUE); c.font = Font(name='Calibri', size=9, color=WHITE, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(LIGHT_BLUE); c.font = font(bold=True, sz=10, color=NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='medium', color=MID_BLUE))
    for ri, row in df_out.iterrows():
        excel_row = ri + 4; ws.row_dimensions[excel_row].height = 13
        row_fill = fill(ALT_ROW) if ri % 2 == 1 else fill(WHITE)
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill = row_fill; c.font = font(sz=9); c.border = thin_border()
            if col in ('Qty Booked','Qty Collected','Ref ID','Date'):
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    total_row = len(df_out) + 4; ws.row_dimensions[total_row].height = 16
    collected_col = headers.index('Qty Collected') + 1
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = fill(TOTAL_BG); c.border = Border(top=Side(style='medium', color=MID_BLUE))
        if ci == 1:
            c.value = 'TOTALS'; c.font = font(bold=True, sz=10, color=NAVY)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci == collected_col:
            c.value = f'=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})'
            c.font = font(bold=True, sz=10, color=NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center')
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(df_out[h], h)
    ws.freeze_panes = 'A4'; ws.auto_filter.ref = f'A3:{get_column_letter(num_cols)}3'

def build_booked(ws, df_data, route_name, date_collected):
    df_out = df_data[list(col_map_booked.keys())].rename(columns=col_map_booked).copy()
    df_out.insert(2, 'Date Collected', date_collected)
    headers = list(df_out.columns); num_cols = len(headers)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    c = ws['A1']; c.value = f'  {route_name}  —  Booked Mattress Collection (WM28)'
    c.fill = fill(ORANGE); c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.date.today().strftime("%d/%m/%Y")}   |   '
               f'Total stops: {len(df_out)}   |   Qty collected: {int(df_out["Qty Collected"].sum())}')
    c.fill = fill(ORANGE_MID); c.font = Font(name='Calibri', size=9, color=WHITE, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(ORANGE_HDR); c.font = font(bold=True, sz=10, color=ORANGE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='medium', color=ORANGE_MID))
    for ri, row in df_out.iterrows():
        excel_row = ri + 4; ws.row_dimensions[excel_row].height = 13
        row_fill = fill(LIGHT_ORANGE) if ri % 2 == 1 else fill(WHITE)
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill = row_fill; c.font = font(sz=9); c.border = thin_border()
            if col in ('Ref ID','Date Booked','Date Collected','Qty Collected'):
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    total_row = len(df_out) + 4; ws.row_dimensions[total_row].height = 16
    collected_col = headers.index('Qty Collected') + 1
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = fill(TOTAL_BG); c.border = Border(top=Side(style='medium', color=ORANGE_MID))
        if ci == 1:
            c.value = 'TOTALS'; c.font = font(bold=True, sz=10, color=ORANGE)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci == collected_col:
            c.value = f'=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})'
            c.font = font(bold=True, sz=10, color=ORANGE)
            c.alignment = Alignment(horizontal='center', vertical='center')
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(df_out[h], h)
    ws.freeze_panes = 'A4'; ws.auto_filter.ref = f'A3:{get_column_letter(num_cols)}3'

def build_burwood(ws, df_data, route_name):
    df_out = df_data[list(col_map_burwood.keys())].rename(columns=col_map_burwood).copy()
    headers = list(col_map_burwood.values()); num_cols = len(headers)
    total_collected = pd.to_numeric(df_out['Collected'], errors='coerce').fillna(0).sum()
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    c = ws['A1']; c.value = f'  {route_name}'
    c.fill = fill(GRN_DARK); c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.date.today().strftime("%d/%m/%Y")}   |   '
               f'Total stops: {len(df_out)}   |   Total collected: {int(total_collected)}')
    c.fill = fill(GRN_MID); c.font = Font(name='Calibri', size=9, color=WHITE, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(GRN_LIGHT); c.font = font(bold=True, sz=10, color=GRN_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='medium', color=GRN_MID))
    for ri, row in df_out.iterrows():
        excel_row = ri + 4; ws.row_dimensions[excel_row].height = 13
        row_fill = fill(GRN_ALT) if ri % 2 == 1 else fill(WHITE)
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill = row_fill; c.font = font(sz=9); c.border = thin_border('C8E6C9')
            if col == 'Collected':
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 'Photo URL':
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                if val and str(val).startswith('http'):
                    first_url = str(val).split(',')[0].strip()
                    c.hyperlink = first_url
                    c.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    total_row = len(df_out) + 4; ws.row_dimensions[total_row].height = 16
    collected_col = headers.index('Collected') + 1
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = fill(TOTAL_BG); c.border = Border(top=Side(style='medium', color=GRN_MID))
        if ci == 1:
            c.value = 'TOTALS'; c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci == collected_col:
            c.value = f'=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})'
            c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = col_width(df_out['Address'], 'Address', cap=60)
    ws.column_dimensions['B'].width = 11
    single_urls = df_out['Photo URL'].astype(str).str.split(',').explode().str.strip()
    ws.column_dimensions['C'].width = min(single_urls.str.len().max() + 2, 160)
    ws.freeze_panes = 'A4'; ws.auto_filter.ref = f'A3:{get_column_letter(num_cols)}3'

def process_randwick(df):
    df = df.drop_duplicates(subset=['seller_order_id', 'address'])
    df['date_booked_dt'] = df['date_booked'].apply(clean_date)
    route_date = df['date_booked_dt'].dropna().mode()[0]
    df['date_booked_dt'] = df['date_booked_dt'].fillna(route_date)
    df['date_booked'] = df['date_booked_dt'].apply(lambda x: x.strftime('%d/%m/%Y'))
    df['suburb'] = df.apply(extract_suburb, axis=1)
    route_name = df['route'].iloc[0]
    route_clean = re.sub(r'^[A-Za-z]+,\s*', '', route_name)
    oncall_filename = f'{route_name} - On Call'
    booked_filename = f'Booked {route_clean}'
    route_date_str = route_date.strftime('%d/%m/%Y')
    df_wm28  = df[df['products'] == 'WM28'].copy()
    df_bulky = df[df['products'] == 'Bulky Mattress'].copy()
    df_blank = df[df['products'].isna() | (df['products'].str.strip() == '')].copy()
    df_wm28  = df_wm28.sort_values(['date_booked_dt','address']).reset_index(drop=True)
    df_bulky = df_bulky.sort_values(['date_booked_dt','suburb','address']).reset_index(drop=True)
    df_blank = df_blank.sort_values(['date_booked_dt','suburb']).reset_index(drop=True)
    df_oncall = pd.concat([df_bulky, df_blank], ignore_index=True)
    wb1 = Workbook(); ws1 = wb1.active; ws1.title = 'On Call'
    build_oncall(ws1, df_oncall, route_name)
    buf1 = io.BytesIO(); wb1.save(buf1); buf1.seek(0)
    wb2 = Workbook(); ws2 = wb2.active; ws2.title = 'Booked'
    build_booked(ws2, df_wm28, route_name, route_date_str)
    buf2 = io.BytesIO(); wb2.save(buf2); buf2.seek(0)
    return [(buf1, f'{oncall_filename}.xlsx', len(df_oncall)),
            (buf2, f'{booked_filename}.xlsx', len(df_wm28))]

col_map_bayside = {
    'address': 'Address',
    'driver_provided_internal_notes': 'Collected',
    'photo_url': 'Photo URL',
    'location_attempted_latitude': 'Latitude',
    'location_attempted_longitude': 'Longitude',
}

def build_bayside(ws, df_data, route_name):
    df_out = df_data[list(col_map_bayside.keys())].rename(columns=col_map_bayside).copy()
    headers = list(col_map_bayside.values()); num_cols = len(headers)
    total_collected = pd.to_numeric(df_out['Collected'], errors='coerce').fillna(0).sum()
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    c = ws['A1']; c.value = f'  {route_name}'
    c.fill = fill(GRN_DARK); c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.date.today().strftime("%d/%m/%Y")}   |   '
               f'Total stops: {len(df_out)}   |   Total collected: {int(total_collected)}')
    c.fill = fill(GRN_MID); c.font = Font(name='Calibri', size=9, color=WHITE, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(GRN_LIGHT); c.font = font(bold=True, sz=10, color=GRN_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='medium', color=GRN_MID))
    for ri, row in df_out.iterrows():
        excel_row = ri + 4; ws.row_dimensions[excel_row].height = 13
        row_fill = fill(GRN_ALT) if ri % 2 == 1 else fill(WHITE)
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill = row_fill; c.font = font(sz=9); c.border = thin_border('C8E6C9')
            if col in ('Collected', 'Latitude', 'Longitude'):
                c.alignment = Alignment(horizontal='center', vertical='center')
            elif col == 'Photo URL':
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                if val and str(val).startswith('http'):
                    first_url = str(val).split(',')[0].strip()
                    c.hyperlink = first_url
                    c.font = Font(name='Calibri', size=9, color='0563C1', underline='single')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    total_row = len(df_out) + 4; ws.row_dimensions[total_row].height = 16
    collected_col = headers.index('Collected') + 1
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = fill(TOTAL_BG); c.border = Border(top=Side(style='medium', color=GRN_MID))
        if ci == 1:
            c.value = 'TOTALS'; c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci == collected_col:
            c.value = f'=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})'
            c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = col_width(df_out['Address'], 'Address', cap=60)
    ws.column_dimensions['B'].width = 11
    single_urls = df_out['Photo URL'].astype(str).str.split(',').explode().str.strip()
    ws.column_dimensions['C'].width = min(single_urls.str.len().max() + 2, 160)
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.freeze_panes = 'A4'; ws.auto_filter.ref = f'A3:{get_column_letter(num_cols)}3'

col_map_innerwest = {
    'address': 'Address',
    'driver_provided_internal_notes': 'Collected',
}

def build_innerwest(ws, df_data, route_name):
    df_out = df_data[list(col_map_innerwest.keys())].rename(columns=col_map_innerwest).copy()
    headers = list(col_map_innerwest.values()); num_cols = len(headers)
    total_collected = pd.to_numeric(df_out['Collected'], errors='coerce').fillna(0).sum()
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
    c = ws['A1']; c.value = f'  {route_name}'
    c.fill = fill(GRN_DARK); c.font = Font(name='Calibri', bold=True, size=14, color=WHITE)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
    c = ws['A2']
    c.value = (f'  Generated: {datetime.date.today().strftime("%d/%m/%Y")}   |   '
               f'Total stops: {len(df_out)}   |   Total collected: {int(total_collected)}')
    c.fill = fill(GRN_MID); c.font = Font(name='Calibri', size=9, color=WHITE, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 18
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.fill = fill(GRN_LIGHT); c.font = font(bold=True, sz=10, color=GRN_DARK)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = Border(bottom=Side(style='medium', color=GRN_MID))
    for ri, row in df_out.iterrows():
        excel_row = ri + 4; ws.row_dimensions[excel_row].height = 13
        row_fill = fill(GRN_ALT) if ri % 2 == 1 else fill(WHITE)
        for ci, (col, val) in enumerate(row.items(), 1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.fill = row_fill; c.font = font(sz=9); c.border = thin_border('C8E6C9')
            if col == 'Collected':
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    total_row = len(df_out) + 4; ws.row_dimensions[total_row].height = 16
    collected_col = headers.index('Collected') + 1
    for ci in range(1, num_cols + 1):
        c = ws.cell(row=total_row, column=ci)
        c.fill = fill(TOTAL_BG); c.border = Border(top=Side(style='medium', color=GRN_MID))
        if ci == 1:
            c.value = 'TOTALS'; c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='left', vertical='center')
        elif ci == collected_col:
            c.value = f'=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})'
            c.font = font(bold=True, sz=10, color=GRN_DARK)
            c.alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['A'].width = col_width(df_out['Address'], 'Address', cap=60)
    ws.column_dimensions['B'].width = 11
    ws.freeze_panes = 'A4'; ws.auto_filter.ref = f'A3:{get_column_letter(num_cols)}3'

def process_burwood(df):
    route_name = df['route'].iloc[0]
    wb = Workbook(); ws = wb.active; ws.title = route_name[:31]
    build_burwood(ws, df, route_name)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return [(buf, f'{route_name}.xlsx', len(df))]

def process_csv(df):
    council = detect_council(df)
    if council == 'randwick': return process_randwick(df), 'randwick'
    elif council == 'burwood':
        df = df.copy()
        df['driver_provided_internal_notes'] = (
            pd.to_numeric(df['driver_provided_internal_notes'], errors='coerce').fillna(0) +
            pd.to_numeric(df['driver_provided_recipient_notes'], errors='coerce').fillna(0)
        )
        return process_burwood(df), 'burwood'
    elif council == 'woollahra':
        df = df.copy()
        df['driver_provided_internal_notes'] = (
            pd.to_numeric(df['driver_provided_internal_notes'], errors='coerce').fillna(0) +
            pd.to_numeric(df['driver_provided_recipient_notes'], errors='coerce').fillna(0)
        )
        return process_burwood(df), 'woollahra'
    elif council == 'penrith':
        df = df.copy()
        df['driver_provided_internal_notes'] = (
            pd.to_numeric(df['driver_provided_internal_notes'], errors='coerce').fillna(0) +
            pd.to_numeric(df['driver_provided_recipient_notes'], errors='coerce').fillna(0)
        )
        route_name = df['route'].iloc[0]
        wb = Workbook(); ws_p = wb.active; ws_p.title = route_name[:31]
        build_innerwest(ws_p, df, route_name)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return [(buf, f'{route_name}.xlsx', len(df))], 'penrith'
    elif council == 'innerwest':
        df = df.copy()
        df['driver_provided_internal_notes'] = (
            pd.to_numeric(df['driver_provided_internal_notes'], errors='coerce').fillna(0) +
            pd.to_numeric(df['driver_provided_recipient_notes'], errors='coerce').fillna(0)
        )
        route_name = df['route'].iloc[0]
        wb = Workbook(); ws_i = wb.active; ws_i.title = route_name[:31]
        build_innerwest(ws_i, df, route_name)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return [(buf, f'{route_name}.xlsx', len(df))], 'innerwest'
    elif council == 'bayside':
        df = df.copy()
        df['driver_provided_internal_notes'] = (
            pd.to_numeric(df['driver_provided_internal_notes'], errors='coerce').fillna(0) +
            pd.to_numeric(df['driver_provided_recipient_notes'], errors='coerce').fillna(0)
        )
        route_name = df['route'].iloc[0]
        wb = Workbook(); ws_b = wb.active; ws_b.title = route_name[:31]
        build_bayside(ws_b, df, route_name)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return [(buf, f'{route_name}.xlsx', len(df))], 'bayside'
    else: raise ValueError(f'Unknown council: {df["route"].iloc[0]}')

# ── UI ────────────────────────────────────────────────────────────────────────
uploaded = st.file_uploader('Upload your route CSV', type=['csv'])

if uploaded:
    df = pd.read_csv(uploaded)
    route_name = df['route'].iloc[0] if 'route' in df.columns else 'Unknown'
    council = detect_council(df)

    col1, col2, col3 = st.columns(3)
    col1.metric('Route', route_name)
    col2.metric('Council', council.title())
    col3.metric('Total Stops', len(df))
    st.divider()

    if st.button('Generate Excel Files'):
        with st.spinner('Processing...'):
            try:
                results, council = process_csv(df)
                st.success('Done! Download your files below.')
                cols = st.columns(len(results))
                for i, (buf, fname, nrows) in enumerate(results):
                    with cols[i]:
                        label = fname.replace('.xlsx','')
                        st.markdown(f"**📋 {label}**")
                        st.download_button(
                            label=f'Download ({nrows} rows)',
                            data=buf,
                            file_name=fname,
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True
                        )
            except Exception as e:
                st.error(f'Error: {e}')
else:
    st.info('👆 Upload a CSV file to get started')
